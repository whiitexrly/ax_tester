import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill

from utils.wcag_helper import WCAG_RULE_MAPPER, get_wcag_level

ISSUE_COLUMNS = [
    "url",
    "wcag_rule",
    "level",
    "description",
    "why it matters",
    "source",
    "html_snippet",
    "fix",
    "image_url_or_path",
]
LEVEL_ORDER = {"A": 0, "AA": 1, "AAA": 2}
REPORT_FILENAME = "ax_report.json"
HEADER_ROW_HEIGHT = 24
DATA_ROW_HEIGHT = 24
WCAG_CHART_DATA_START_COL = 26
POUR_PRINCIPLES = {
    "1": "1 - Perceivable",
    "2": "2 - Operable",
    "3": "3 - Understandable",
    "4": "4 - Robust",
}


def build_excel_report(results_dir: str) -> str:
    if not results_dir:
        raise ValueError("results_dir not found in tool_context.state. Run save first.")

    report = _load_report(results_dir)
    all_issues = _load_issues_with_page_url(results_dir, report)
    all_issues_df = _build_all_issues_df(all_issues)
    wcag_summary_df = _build_wcag_summary(all_issues_df)
    level_summary_df = _build_level_summary(all_issues)

    out_xlsx = str(Path(results_dir) / "ax_report.xlsx")
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        level_summary_df.to_excel(writer, sheet_name="Level Summary", index=False, header=False)
        wcag_summary_df.to_excel(writer, sheet_name="WCAG Summary", index=False)
        all_issues_df.to_excel(writer, sheet_name="All Issues", index=False)
        _format_workbook(writer)

    return out_xlsx


def _build_all_issues_df(all_issues: list[dict[str, Any]]) -> pd.DataFrame:
    df = pd.DataFrame(all_issues)
    if df.empty:
        return pd.DataFrame(columns=ISSUE_COLUMNS)

    df["level"] = df.get("wcag_rule").apply(get_wcag_level)
    df["why it matters"] = df.get("why_this_matters")
    df["url"] = df.apply(_get_issue_url, axis=1)
    for column in ISSUE_COLUMNS:
        if column not in df.columns:
            df[column] = None
    df = df[ISSUE_COLUMNS]
    return _sort_with_level(df, by=["wcag_rule", "source", "url"], ascending=[True, True, True])


def _build_level_summary(all_issues: list[dict[str, Any]]) -> pd.DataFrame:
    level_counts = {"A": 0, "AA": 0, "AAA": 0}
    for issue in all_issues:
        level_counts[get_wcag_level(issue.get("wcag_rule"))] += 1

    return pd.DataFrame(
        {
            "summary": [
                f"Level A issues: {level_counts['A']}",
                f"Level AA issues: {level_counts['AA']}",
                f"Level AAA issues: {level_counts['AAA']}",
            ]
        }
    )


def _build_wcag_summary(all_issues_df: pd.DataFrame) -> pd.DataFrame:
    baseline_df = pd.DataFrame({"wcag_rule": _get_a_aa_wcag_rules()})
    if baseline_df.empty:
        return pd.DataFrame(columns=["wcag_rule", "level", "total_issues"])

    if all_issues_df.empty:
        summary_df = baseline_df.assign(total_issues=0)
    else:
        observed_counts_df = (
            all_issues_df.groupby("wcag_rule", dropna=False).size().reset_index(name="total_issues")
        )
        summary_df = baseline_df.merge(observed_counts_df, on="wcag_rule", how="left")
        summary_df["total_issues"] = summary_df["total_issues"].fillna(0).astype(int)

        extra_rules_df = observed_counts_df[~observed_counts_df["wcag_rule"].isin(summary_df["wcag_rule"])]
        if not extra_rules_df.empty:
            summary_df = pd.concat([summary_df, extra_rules_df], ignore_index=True)

    summary_df["level"] = summary_df["wcag_rule"].apply(get_wcag_level)
    summary_df = _sort_with_level(summary_df, by=["total_issues", "wcag_rule"], ascending=[False, True])
    return summary_df[["wcag_rule", "level", "total_issues"]]


def _sort_with_level(df: pd.DataFrame, by: list[str], ascending: list[bool]) -> pd.DataFrame:
    if df.empty:
        return df

    sorted_df = df.assign(level_order=df["level"].map(_level_order))
    sorted_df = sorted_df.sort_values(
        by=["level_order", *by],
        ascending=[True, *ascending],
        kind="stable",
    )
    return sorted_df.drop(columns=["level_order"])


def _get_a_aa_wcag_rules() -> list[str]:
    return list(dict.fromkeys(rule for rule in WCAG_RULE_MAPPER.values() if get_wcag_level(rule) in {"A", "AA"}))


def _level_order(level: object) -> int:
    if isinstance(level, str):
        return LEVEL_ORDER.get(level.upper(), 3)
    return 3


def _safe_int(value: object) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _format_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    workbook["Level Summary"].column_dimensions["A"].width = 40

    for sheet_name in ["WCAG Summary", "All Issues"]:
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

        for row_index, row in enumerate(ws.iter_rows(min_row=2), start=1):
            row_fill = odd_row_fill if row_index % 2 else even_row_fill
            ws.row_dimensions[row_index + 1].height = DATA_ROW_HEIGHT
            for cell in row:
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="center")

        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = max(len("" if cell.value is None else str(cell.value)) for cell in col_cells)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 70)

        if sheet_name == "WCAG Summary":
            _add_wcag_summary_chart(ws, header_fill, header_font, odd_row_fill, even_row_fill)


def _add_wcag_summary_chart(
    ws: Any,
    header_fill: PatternFill,
    header_font: Font,
    odd_row_fill: PatternFill,
    even_row_fill: PatternFill,
) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}
    wcag_rule_col = headers.get("wcag_rule")
    total_issues_col = headers.get("total_issues")
    if wcag_rule_col is None or total_issues_col is None:
        return

    issue_counts_by_principle = dict.fromkeys([*POUR_PRINCIPLES.values(), "Other"], 0)
    for row_index in range(2, ws.max_row + 1):
        wcag_rule = ws.cell(row=row_index, column=wcag_rule_col).value
        total_issues = _safe_int(ws.cell(row=row_index, column=total_issues_col).value)
        if not wcag_rule or total_issues <= 0:
            continue
        issue_counts_by_principle[_pour_principle_label(wcag_rule)] += total_issues

    chart_rows = [(principle, count) for principle, count in issue_counts_by_principle.items() if count > 0]
    if not chart_rows:
        return

    label_col = WCAG_CHART_DATA_START_COL
    value_col = WCAG_CHART_DATA_START_COL + 1
    ws.cell(row=1, column=label_col, value="POUR principle")
    ws.cell(row=1, column=value_col, value="Total issues")
    for cell in (ws.cell(row=1, column=label_col), ws.cell(row=1, column=value_col)):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center")

    for row_index, (principle, total_issues) in enumerate(chart_rows, start=2):
        row_fill = odd_row_fill if row_index % 2 else even_row_fill
        label_cell = ws.cell(row=row_index, column=label_col, value=principle)
        value_cell = ws.cell(row=row_index, column=value_col, value=total_issues)
        for cell in (label_cell, value_cell):
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")

    chart = PieChart()
    chart.style = 10
    chart.title = "Issues by POUR principle"
    chart.height = 9
    chart.width = 14
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = True
    chart.dataLabels.showLeaderLines = True

    data = Reference(ws, min_col=value_col, min_row=1, max_row=len(chart_rows) + 1)
    categories = Reference(ws, min_col=label_col, min_row=2, max_row=len(chart_rows) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.visible_cells_only = False
    ws.add_chart(chart, "E2")

    # Keep source columns visible: Excel may render charts blank when their
    # source range is hidden.
    ws.column_dimensions["Z"].width = 45
    ws.column_dimensions["AA"].width = 14


def _pour_principle_label(wcag_rule: object) -> str:
    rule = str(wcag_rule or "").strip()
    if not rule:
        return "Other"
    return POUR_PRINCIPLES.get(rule[0], "Other")


def _load_report(results_dir: str) -> dict[str, Any]:
    report_path = Path(results_dir) / REPORT_FILENAME
    if not report_path.exists():
        raise FileNotFoundError(f"Missing report file: {report_path}")

    with open(report_path, encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise ValueError(f"Invalid report format in {report_path}: expected JSON object")
    return data


def _load_issues_with_page_url(results_dir: str, report: dict[str, Any]) -> list[dict[str, Any]]:
    results_path = _find_results_index(Path(results_dir))
    if not results_path.exists():
        return _extract_issue_list_with_page_url(report)

    try:
        with open(results_path, encoding="utf-8") as f:
            reports = json.load(f)
    except Exception:
        return _extract_issue_list_with_page_url(report)

    if not isinstance(reports, list):
        return _extract_issue_list_with_page_url(report)

    all_issues: list[dict[str, Any]] = []
    for page_report in reports:
        if not isinstance(page_report, dict):
            continue
        all_issues.extend(_extract_issue_list_with_page_url(page_report))
    return all_issues


def _find_results_index(results_dir: Path) -> Path:
    for filename in ("results.json", "results_bpm.json"):
        candidate = results_dir / filename
        if candidate.exists():
            return candidate
    return results_dir / "results.json"


def _extract_issue_list_with_page_url(report: dict[str, Any]) -> list[dict[str, Any]]:
    page_url = str(report.get("page") or "").strip()
    issues = []
    for issue in _extract_issue_list(report):
        issue_with_url = issue.copy()
        if page_url and not _get_issue_url(issue_with_url):
            issue_with_url["url"] = page_url
        issues.append(issue_with_url)
    return issues


def _get_issue_url(issue: Any) -> str:
    if hasattr(issue, "get"):
        for key in ("url", "page", "page_url", "_report_page"):
            value = str(issue.get(key) or "").strip()
            if value:
                return value
    return ""


def _extract_issue_list(report: dict[str, Any]) -> list[dict[str, Any]]:
    issue_list = report.get("issue_list", [])
    all_issues: list[dict[str, Any]] = []
    if not isinstance(issue_list, list):
        return all_issues

    for issue in issue_list:
        if isinstance(issue, dict):
            all_issues.append(issue)
    return all_issues


if __name__ == "__main__":
    results_dir = "/home/pbianco/ax_tester/results/2026-05-18_15-54-29_shop.reply.com/2026-05-18_15-58-56"
    build_excel_report(results_dir)
